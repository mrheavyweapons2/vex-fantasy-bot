"""
File: manager/excel.py
Author: Jeremiah Nairn

Description: Holds all of the functionality for making excel files and displaying them as images
"""

#import libraries
import openpyxl as xl
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

#class to manage the excel file
class ExcelManager:

    #constructor
    def __init__(self,filename,playerdata,rounds,total_players):
        #important variables
        self.workbook = xl.Workbook()
        self.sheet = self.workbook.active
        self.filename = filename
        self.rounds = rounds
        self.draft_data = playerdata
        self.total_players = total_players
        #create two sheets, one for the draft, and one for the results
        self.sheet.title = "Draft"
        self.workbook.create_sheet(title="Results")
        #define styles
        self.header_fill = xl.styles.PatternFill("solid", fgColor="0F243E")  #dark blue
        self.header_font = xl.styles.Font(color="FFFFFF", bold=True, size=12)
        self.body_fill = xl.styles.PatternFill("solid", fgColor="0c0a0f")  #almost black
        self.body_font = xl.styles.Font(color="FFFFFF", bold=False, size=12)
        self.alignment = xl.styles.Alignment(horizontal="center", vertical="center")
        self.all_border = xl.styles.Border(
            left=xl.styles.Side(style="thin",color="FFFFFF"), right=xl.styles.Side(style="thin",color="FFFFFF"),
            top=xl.styles.Side(style="thin",color="FFFFFF"), bottom=xl.styles.Side(style="thin",color="FFFFFF"),
        )
        self.side_borders = xl.styles.Border(
            left=xl.styles.Side(style="thin",color="FFFFFF"), right=xl.styles.Side(style="thin",color="FFFFFF")
        )
        return

    #function to save the excel file
    def save_excel(self):
        #locate the excels folder
        if not os.path.exists('excels'):
            os.makedirs('excels')
        #save the file
        self.workbook.save(f"excels/{self.filename}.xlsx")
        print(f"[EXCEL] Excel file saved as {f"{self.filename}.xlsx"}")

    def update_playerdata(self,playerdata):
        #clear the current playerdata
        self.draft_data = playerdata

    #function to create the draft sheet
    def create_draft_sheet(self):
        #create the header
        self.sheet["A1"] = "Drafter"
        self.sheet["A1"].font = self.header_font
        self.sheet["A1"].fill = self.header_fill
        self.sheet["A1"].border = self.all_border
        self.sheet["A1"].alignment = self.alignment
        self.sheet.column_dimensions['A'].width = 30
        #create the round headers
        for round in range(self.rounds):
            col = chr(66 + round)  # ASCII value for 'B' is 66
            cell = f"{col}1"
            self.sheet[cell] = f"Pick {round + 1}"
            self.sheet[cell].font = self.header_font
            self.sheet[cell].fill = self.header_fill
            self.sheet[cell].border = self.all_border
            self.sheet[cell].alignment = self.alignment
            self.sheet.column_dimensions[col].width = 15
        #get the position of each player, and format each collumn
        for player in range(self.total_players):
            row = player + 2  # Starting from row 2 since row 1 is the header
            drafter_cell = f"A{row}"
            #identify the player by their draft position
            for drafter in self.draft_data:
                if drafter['position'] == player+1:
                    self.sheet[drafter_cell] = drafter['name']
            #format the cells in the row
            for round in range(self.rounds+1):
                col = chr(65 + round)
                pick_cell = f"{col}{row}"
                self.sheet[pick_cell].border = self.side_borders
                self.sheet[pick_cell].alignment = self.alignment
                self.sheet[pick_cell].font = self.body_font
                self.sheet[pick_cell].fill = self.body_fill

        #save the file and return
        self.save_excel()
        return

    #function to fill in the draft sheet with data
    def fill_draft_sheet(self, draft_data):
        #get the position of each player, and format each collumn
        for player in range(self.total_players):
            row = player + 2  # Starting from row 2 since row 1 is the header
            #identify the player by their draft position
            for drafter in self.draft_data:
                if drafter['position'] == player+1:
                    #format the cells in the row
                    for round in range(self.rounds):
                        col = chr(66 + round)
                        pick_cell = f"{col}{row}"
                        #get the pick and put it in the cell
                        self.sheet[pick_cell] = drafter[f"round_{round+1}"]
        #save the file and return
        self.save_excel()
        return

    #function to get the draft as an image (no win32)
    def get_draft_as_image(self):
        # Save workbook to ensure on-disk copy
        self.save_excel()
        # load workbook from saved file to get final cell values/styles
        wb_path = os.path.abspath(f"excels/{self.filename}.xlsx")
        wb = load_workbook(wb_path)
        ws = wb["Draft"]
        # determine bounds
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1
        # helpers for converting sizes/colors
        def col_width_to_px(width):
            # approximate conversion (Excel char width -> px)
            if width is None:
                width = 8.43
            return int(width * 7 + 5)
        def row_height_to_px(height):
            # height in points -> pixels (96dpi)
            if height is None:
                height = 15
            return int(height * 96 / 72)
        def parse_color(c):
            # c is a Color object; attempt to get an rgb hex and return (r,g,b)
            if c is None:
                return (255, 255, 255)
            # many Color objects expose .rgb as "FFRRGGBB" or "RRGGBB"
            val = None
            try:
                val = getattr(c, "rgb", None)
            except Exception:
                val = None
            if not val:
                try:
                    val = getattr(c, "fgColor", None)
                    if val is not None:
                        val = getattr(val, "rgb", None)
                except Exception:
                    val = None
            if not val:
                # fallback white
                return (255, 255, 255)
            s = str(val)
            s = s.lstrip("#")
            if len(s) == 8:  # maybe "FFRRGGBB"
                s = s[2:]
            if len(s) != 6:
                return (255, 255, 255)
            try:
                return tuple(int(s[i : i + 2], 16) for i in (0, 2, 4))
            except Exception:
                return (255, 255, 255)
        # compute pixel sizes for columns and rows
        col_pixels = []
        for c in range(1, max_col + 1):
            letter = get_column_letter(c)
            cd = ws.column_dimensions.get(letter)
            w = None
            if cd is not None:
                w = getattr(cd, "width", None)
            col_pixels.append(col_width_to_px(w))
        row_pixels = []
        for r in range(1, max_row + 1):
            rd = ws.row_dimensions.get(r)
            h = None
            if rd is not None:
                h = getattr(rd, "height", None)
            row_pixels.append(row_height_to_px(h))
        # layout parameters
        border_px = 1
        pad_x = 8
        pad_y = 6
        total_width = sum(col_pixels) + (max_col + 1) * border_px
        total_height = sum(row_pixels) + (max_row + 1) * border_px
        # create image
        img = Image.new("RGB", (total_width + 2 * pad_x, total_height + 2 * pad_y), (255, 255, 255))
        draw = ImageDraw.Draw(img)
        # try to load a reasonable truetype font, fallback to default
        def get_font(size, bold=False):
            try:
                # common Windows font
                return ImageFont.truetype("arialbd.ttf" if bold else "arial.ttf", size)
            except Exception:
                try:
                    return ImageFont.truetype("DejaVuSans-Bold.ttf" if bold else "DejaVuSans.ttf", size)
                except Exception:
                    return ImageFont.load_default()
        # draw cells
        y_offset = pad_y
        for r_idx in range(max_row):
            x_offset = pad_x
            row_h = row_pixels[r_idx]
            for c_idx in range(max_col):
                col_w = col_pixels[c_idx]
                left = x_offset
                top = y_offset
                right = left + col_w
                bottom = top + row_h

                cell = ws.cell(row=r_idx + 1, column=c_idx + 1)
                # background fill
                bg_color = (255, 255, 255)
                try:
                    if cell.fill and getattr(cell.fill, "fgColor", None) is not None:
                        bg_color = parse_color(cell.fill.fgColor)
                except Exception:
                    bg_color = (255, 255, 255)
                draw.rectangle([left, top, right, bottom], fill=bg_color)

                # border (simple)
                draw.rectangle([left, top, right, bottom], outline=(200, 200, 200), width=border_px)

                # text
                value = "" if cell.value is None else str(cell.value)
                # font size from cell.font.sz (points) fallback 12
                fsize = 12
                bold = False
                try:
                    if cell.font is not None:
                        if getattr(cell.font, "sz", None):
                            fsize = int(getattr(cell.font, "sz"))
                        bold = bool(getattr(cell.font, "bold", False))
                except Exception:
                    pass
                font = get_font(max(8, fsize), bold=bold)

                # font color
                font_color = (0, 0, 0)
                try:
                    if cell.font is not None and getattr(cell.font, "color", None) is not None:
                        fc = getattr(cell.font.color, "rgb", None)
                        if fc:
                            font_color = parse_color(cell.font.color)
                except Exception:
                    font_color = (0, 0, 0)

                # alignment
                halign = "center"
                valign = "center"
                try:
                    if cell.alignment is not None:
                        halign = getattr(cell.alignment, "horizontal", halign) or halign
                        valign = getattr(cell.alignment, "vertical", valign) or valign
                except Exception:
                    pass

                #measure and position text
                #measure and position text
                try:
                    # Preferred modern API: returns (left, top, right, bottom)
                    bbox = draw.textbbox((0, 0), value, font=font)
                    txt_w = bbox[2] - bbox[0]
                    txt_h = bbox[3] - bbox[1]
                except Exception:
                    # Fallback: try font.getsize, or default to zero-sized text
                    try:
                        txt_w, txt_h = font.getsize(value)
                    except Exception:
                        txt_w, txt_h = (0, 0)

                if halign == "left":
                    tx = left + 4
                elif halign == "right":
                    tx = right - txt_w - 4
                else:
                    tx = left + (col_w - txt_w) / 2

                if valign == "top":
                    ty = top + 2
                elif valign == "bottom":
                    ty = bottom - txt_h - 2
                else:
                    ty = top + (row_h - txt_h) / 2

                draw.text((tx, ty), value, font=font, fill=font_color)
                x_offset += col_w + border_px
            y_offset += row_h + border_px

        # ensure directory exists and save
        base_dir = os.path.dirname(os.path.abspath(__file__))  # manager folder
        export_dir = os.path.join(base_dir, "excels")
        os.makedirs(export_dir, exist_ok=True)
        image_path = os.path.join(export_dir, f"{self.filename}_draft.png")
        img.save(image_path, format="PNG")

        return image_path

        

    #function to create the results sheet
    def create_results_sheet(self):

        #save the file and return
        self.save_excel()
        return

    #function to fill in the results sheet with data
    def fill_results_sheet(self, draft_data, results_data):

        #save the file and return
        self.save_excel()
        return

    #function to get the results
    def get_results_as_image(self):
        return

#function to wipe the excel folder
def wipe_excel_folder():
    #delete all xlsx files in the excels folder
    for file in os.listdir('excels'):
        if file.endswith('.xlsx'):
            os.remove(os.path.join('excels', file))
    return
